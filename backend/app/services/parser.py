"""Document parsers: returns list of (text, page_number) tuples."""
import csv
import io
from pathlib import Path

import fitz  # PyMuPDF
import markdown
from bs4 import BeautifulSoup
from docx import Document
from openpyxl import load_workbook


def parse_pdf(file_bytes: bytes) -> list[tuple[str, int]]:
    doc = fitz.open(stream=file_bytes, filetype="pdf")
    try:
        pages = []
        for page_num, page in enumerate(doc, start=1):
            text = page.get_text("text")
            if text.strip():
                pages.append((text, page_num))
        return pages
    finally:
        doc.close()


def parse_docx(file_bytes: bytes) -> list[tuple[str, int]]:
    doc = Document(io.BytesIO(file_bytes))
    paragraphs = [p.text for p in doc.paragraphs if p.text.strip()]
    # DOCX has no page numbers; group all text as page 1
    text = "\n".join(paragraphs)
    return [(text, 1)] if text else []


def parse_html(file_bytes: bytes) -> list[tuple[str, int]]:
    soup = BeautifulSoup(file_bytes, "lxml")
    # Remove scripts and styles
    for tag in soup(["script", "style", "head"]):
        tag.decompose()
    text = soup.get_text(separator="\n", strip=True)
    return [(text, 1)] if text else []


def parse_text(file_bytes: bytes) -> list[tuple[str, int]]:
    text = file_bytes.decode("utf-8", errors="replace")
    return [(text, 1)] if text.strip() else []


def parse_markdown(file_bytes: bytes) -> list[tuple[str, int]]:
    md_text = file_bytes.decode("utf-8", errors="replace")
    html = markdown.markdown(md_text)
    soup = BeautifulSoup(html, "lxml")
    text = soup.get_text(separator="\n", strip=True)
    return [(text, 1)] if text else []


def parse_csv(file_bytes: bytes) -> list[tuple[str, int]]:
    text = file_bytes.decode("utf-8", errors="replace")
    reader = csv.reader(io.StringIO(text))
    rows = list(reader)
    if not rows:
        return []
    headers = rows[0]
    # Convert each row into a readable "Header: Value" representation
    lines: list[str] = []
    for row in rows[1:]:
        parts = [f"{h}: {v}" for h, v in zip(headers, row) if v.strip()]
        if parts:
            lines.append(" | ".join(parts))
    if not lines:
        # Fallback: just join all rows as plain text
        lines = [", ".join(r) for r in rows if any(c.strip() for c in r)]
    result = "\n".join(lines)
    return [(result, 1)] if result else []


def parse_xlsx(file_bytes: bytes) -> list[tuple[str, int]]:
    wb = load_workbook(filename=io.BytesIO(file_bytes), read_only=True, data_only=True)
    pages: list[tuple[str, int]] = []
    for sheet_num, sheet in enumerate(wb.worksheets, start=1):
        rows = list(sheet.iter_rows(values_only=True))
        if not rows:
            continue
        headers = [str(c) if c is not None else "" for c in rows[0]]
        lines: list[str] = []
        for row in rows[1:]:
            cells = [str(c) if c is not None else "" for c in row]
            parts = [f"{h}: {v}" for h, v in zip(headers, cells) if v.strip()]
            if parts:
                lines.append(" | ".join(parts))
        if not lines:
            lines = [
                ", ".join(str(c) for c in row if c is not None)
                for row in rows
                if any(c is not None for c in row)
            ]
        text = f"Sheet: {sheet.title}\n" + "\n".join(lines)
        pages.append((text, sheet_num))
    wb.close()
    return pages


def parse_document(file_bytes: bytes, filename: str) -> list[tuple[str, int]]:
    """Route to the correct parser based on file extension."""
    suffix = Path(filename).suffix.lower()
    if suffix == ".pdf":
        return parse_pdf(file_bytes)
    elif suffix == ".docx":
        return parse_docx(file_bytes)
    elif suffix in (".html", ".htm"):
        return parse_html(file_bytes)
    elif suffix == ".md":
        return parse_markdown(file_bytes)
    elif suffix == ".csv":
        return parse_csv(file_bytes)
    elif suffix == ".xlsx":
        return parse_xlsx(file_bytes)
    else:  # .txt and anything else
        return parse_text(file_bytes)
